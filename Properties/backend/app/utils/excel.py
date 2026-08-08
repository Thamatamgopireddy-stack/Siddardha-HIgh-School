import io
import csv
from datetime import date, datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _format_cell_value(cell_val: Any) -> str:
    if cell_val is None:
        return ""
    if isinstance(cell_val, (datetime, date)):
        return cell_val.strftime("%Y-%m-%d")
    if isinstance(cell_val, float):
        if cell_val.is_integer():
            return str(int(cell_val))
        return f"{cell_val:f}".rstrip('0').rstrip('.')
    return str(cell_val).strip()

def _norm_str(val: Any) -> str:
    return "".join(ch.lower() for ch in str(val or "") if ch.isalnum())

def parse_excel_or_csv(file_bytes: bytes, filename: str) -> List[Dict[str, str]]:
    """
    Parses file bytes of .xlsx, .xls, or .csv file across all worksheets.
    Dynamically detects header row in each sheet and returns a consolidated list
    of dictionaries containing clean string representations of data.
    """
    filename_lower = filename.lower()
    all_rows: List[Dict[str, str]] = []

    if filename_lower.endswith((".xlsx", ".xls")):
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]
                if sheet is None or sheet.max_row < 1:
                    continue

                # Find header row dynamically (look for row containing name/student/roll/adm/id)
                header_row_idx: Optional[int] = None
                headers: List[str] = []

                max_scan_rows = min(sheet.max_row, 15)
                for r_idx in range(1, max_scan_rows + 1):
                    row_cells = list(sheet[r_idx])
                    row_vals = [_format_cell_value(c.value) for c in row_cells]
                    norm_vals = [_norm_str(v) for v in row_vals]

                    # Header detection heuristics
                    if any(kw in nv for nv in norm_vals for kw in ["name", "student", "roll", "adm", "surname", "pen"]):
                        header_row_idx = r_idx
                        headers = row_vals
                        break

                if header_row_idx is None:
                    # Fallback to row 1 if no explicit header row detected
                    header_row_idx = 1
                    headers = [_format_cell_value(c.value) for c in list(sheet[1])]

                # Extract data rows starting after header row
                for row_tuple in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    if not row_tuple or not any(row_tuple):
                        continue
                    row_dict: Dict[str, str] = {}
                    for idx, cell_val in enumerate(row_tuple):
                        if idx < len(headers) and headers[idx] and headers[idx].lower() != 'none':
                            formatted_val = _format_cell_value(cell_val)
                            if formatted_val and formatted_val.lower() != 'none':
                                row_dict[headers[idx]] = formatted_val
                    
                    if any(row_dict.values()):
                        # Add sheet name context if helpful
                        row_dict["_sheet_name"] = sheetname
                        all_rows.append(row_dict)

            if all_rows:
                return all_rows
        except Exception:
            # Fallback to CSV parsing if Excel workbook reading fails
            pass

    # CSV Parsing fallback
    try:
        decoded = file_bytes.decode("utf-8-sig", errors="ignore")
    except Exception:
        decoded = file_bytes.decode("latin-1", errors="ignore")

    reader = csv.DictReader(io.StringIO(decoded))
    for row in reader:
        cleaned_row = {str(k).strip(): str(v).strip() if v is not None else "" for k, v in row.items() if k}
        if any(cleaned_row.values()):
            all_rows.append(cleaned_row)

    return all_rows


def generate_excel_template(
    headers: List[str],
    sample_rows: Optional[List[Dict[str, str]]] = None,
    sheet_name: str = "Template"
) -> bytes:
    """
    Generates a formatted .xlsx template with styled header row and optional sample rows.
    Returns file bytes ready to be streamed as a file response.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Active worksheet is None")
    ws.title = sheet_name

    # Header styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark slate
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.append(headers)

    # Apply header styles
    ws.row_dimensions[1].height = 28
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add sample rows if provided
    if sample_rows:
        for row_data in sample_rows:
            row_vals = [row_data.get(h, "") for h in headers]
            ws.append(row_vals)

    # Style data rows and auto-fit column widths
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column or 1)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


